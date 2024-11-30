import pandas as pd
from fpdf import FPDF
from openpyxl import Workbook
import os

def generate_report(participant_name: str, data: pd.DataFrame, model_path: str) -> str:
    """
    Generiert einen Bericht für einen Teilnehmer im PDF- und Excel-Format.

    Args:
        participant_name (str): Der Name des Teilnehmers.
        data (pd.DataFrame): Der vollständige Datensatz.
        model_path (str): Der Pfad zum gespeicherten H2O-Modell.

    Returns:
        str: Der Pfad zum PDF-Bericht.
    """
    # Teilnehmerdaten extrahieren
    participant_data = data[data["Name"] == participant_name]
    if participant_data.empty:
        raise ValueError(f"Keine Daten für Teilnehmer: {participant_name}")

    # Historische Daten und Vorhersagen
    from src.utils.predictions import generate_h2o_predictions
    predictions = generate_h2o_predictions(participant_data, model_path)

    # Berichtspfad
    report_dir = "reports"
    os.makedirs(report_dir, exist_ok=True)
    pdf_path = os.path.join(report_dir, f"{participant_name}-Bericht.pdf")
    excel_path = os.path.join(report_dir, f"{participant_name}-Bericht.xlsx")

    # PDF-Bericht generieren
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    pdf.cell(200, 10, txt=f"Bericht für {participant_name}", ln=True, align="C")
    pdf.ln(10)

    pdf.cell(200, 10, txt="Teilnehmerdetails:", ln=True)
    for col in participant_data.columns:
        pdf.cell(200, 10, txt=f"{col}: {participant_data[col].values[0]}", ln=True)

    pdf.ln(10)
    pdf.cell(200, 10, txt="Vorhersagedaten:", ln=True)
    for idx, row in predictions.iterrows():
        pdf.cell(200, 10, txt=f"Tag {idx + 1}: {row[0]}%", ln=True)

    pdf.output(pdf_path)

    # Excel-Bericht generieren
    wb = Workbook()
    ws = wb.active
    ws.title = "Bericht"

    # Teilnehmerdetails
    ws.append(["Teilnehmerdetails"])
    for col in participant_data.columns:
        ws.append([col, participant_data[col].values[0]])

    ws.append([])
    ws.append(["Vorhersagedaten"])
    for idx, row in predictions.iterrows():
        ws.append([f"Tag {idx + 1}", row[0]])

    wb.save(excel_path)

    return pdf_path
